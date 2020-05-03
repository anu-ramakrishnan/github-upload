import pandas as pd
import openpyxl
import random
from collections import Counter

#File to be copied
mentees = openpyxl.load_workbook("mentees.xlsx") #Add file name
mentees_sheet = mentees.get_sheet_by_name("Form1") #Add Sheet name

mentors = openpyxl.load_workbook("mentors.xlsx") #Add file name
mentors_sheet = mentors.get_sheet_by_name("Form1") #Add Sheet name

cclist = openpyxl.load_workbook("Consolidated candidates list.xlsx") #Add file name
tees_sheet = cclist.get_sheet_by_name("Mentees (D) - Female") #Add Sheet name
edonly_sheet = cclist.get_sheet_by_name("Mentees+Mentors (ED) - Female") #Add Sheet name
tors_sheet = cclist.get_sheet_by_name("Mentors(ED+ Male, VP+ Female)") #Add Sheet name


#2018 data
list2018 = openpyxl.load_workbook("2018list.xlsx") #Add file name
sheet2018 = list2018.get_sheet_by_name("Mentees matched to mentors") #Add Sheet name
 
#File to be pasted into
matched = openpyxl.load_workbook("matched_listc.xlsx") #Add file name
matched_sheet = matched.get_sheet_by_name("Sheet1") #Add Sheet name
mentor_view = matched.get_sheet_by_name("Sheet2")

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

#global variables to hold the r1,r2 and r3 evolving lists
aggregatedmentorlist = []
aggregated2018list = [[]]

def canMentorbeused(mentee_name, mentor_name):
    
    #check 2018 list
    for i in range(len(aggregated2018list)):
        if(mentee_name == aggregated2018list[i][0]):
            if(mentor_name == aggregated2018list[i][1] or mentor_name == aggregated2018list[i][2]):
                return 0
        
    #use the mentor for upto 4 mentees
#    print(mentor_name,aggregatedmentorlist.count(mentor_name))
    if(aggregatedmentorlist.count(mentor_name) <= 3):
        return 1
    else:
        return 0
    
       
def createData():
    print("Processing...")
    
    submentorlist = {}
    list2018mentees = []
    
    submentorlist1 = {}
    
    # copy name
    selectedRange = copyRange(5,2,5,48,mentees_sheet) 
    pasteRange(2,2,2,48,matched_sheet,selectedRange) 
    
    # copy email
    selectedRange = copyRange(4,2,4,48,mentees_sheet) 
    pasteRange(1,2,1,48,matched_sheet,selectedRange) 
    
    # copy cohort
    selectedRange = copyRange(9,2,9,48,mentees_sheet)
    pasteRange(5,2,5,48,matched_sheet,selectedRange)
    
    #copy mentors to mentors view
    selectedRange = copyRange(5,2,5,48,mentors_sheet) 
    pasteRange(1,2,1,48,mentor_view,selectedRange)
    
    # copy department - after matching
    for matched_name in range(2,matched_sheet.max_row+1):
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == tees_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,5).value = tees_sheet.cell(cc_name,9).value
                
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == edonly_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,5).value = edonly_sheet.cell(cc_name,9).value
                 
    # copy company - after matching
    for matched_name in range(2,matched_sheet.max_row+1):
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == tees_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,3).value = tees_sheet.cell(cc_name,6).value
                
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == edonly_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,3).value = edonly_sheet.cell(cc_name,6).value
                 
    # copy location - after matching
    for matched_name in range(2,matched_sheet.max_row+1):
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == tees_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,4).value = tees_sheet.cell(cc_name,10).value
                
        for cc_name in range(2,tees_sheet.max_row+1):            
            if matched_sheet.cell(matched_name,2).value == edonly_sheet.cell(cc_name,3).value:
                matched_sheet.cell(matched_name,4).value = edonly_sheet.cell(cc_name,10).value       
                 
    
#    for matched_name in range(2,mentors_sheet.max_row+1):
#        for cc_name in range(2,tors_sheet.max_row+1):            
#            if mentors_sheet.cell(matched_name,5).value == tors_sheet.cell(cc_name,3).value:
#                mentors_sheet.cell(matched_name,10).value = tors_sheet.cell(cc_name,9).value
#                
#        for cc_name in range(2,mentees_sheet.max_row+1):            
#            if mentors_sheet.cell(matched_name,5).value == edonly_sheet.cell(cc_name,3).value:
#                mentors_sheet.cell(matched_name,10).value = edonly_sheet.cell(cc_name,9).value
#    
#    mentors.save("mentors.xlsx")    
    
    
    
    #extract 2018 data
    for mentee_name in range(2,matched_sheet.max_row+1):
        for ee_name in range(2,sheet2018.max_row+1):
            if matched_sheet.cell(mentee_name,2).value == sheet2018.cell(ee_name,3).value:
                list2018mentees = []
                list2018mentees.append(matched_sheet.cell(mentee_name,2).value)
                list2018mentees.append(sheet2018.cell(ee_name,10).value)
                list2018mentees.append(sheet2018.cell(ee_name,11).value)
                list2018mentees.append(sheet2018.cell(ee_name,12).value)
                aggregated2018list.append(list2018mentees)                 

    aggregated2018list.pop(0)
#loop through the list of mentees
#for each mentee, find 3 mentors --
#    loop through the mentor list and apply business rules to find the sub-set of mentors
#    use rand to find the first mentor and update the sub-set list to remove that mentor
#    use rand to find the second mentor on this sub-set list and update the sub-set list to remove that mentor
#    use rand to find the third mentor from this sub-set list.
  
    for mentee_name in range(2,matched_sheet.max_row+1):
        submentorlist = {}
        for mentor in range(2,mentors_sheet.max_row+1):
            #cohort, 1 mentor - 3 mentees & no 2018 matching
            if(matched_sheet.cell(mentee_name,6).value != mentors_sheet.cell(mentor,9).value):
                if canMentorbeused(matched_sheet.cell(mentee_name,2).value, mentors_sheet.cell(mentor,5).value):
                    if(mentors_sheet.cell(mentor,5).value not in submentorlist):
#                        submentorlist.append(mentors_sheet.cell(mentor,5).value)
                        submentorlist[(mentors_sheet.cell(mentor,5).value)] = aggregatedmentorlist.count(mentors_sheet.cell(mentor,5).value)

        temp =  min(submentorlist.values())
        res = [key for key in submentorlist if submentorlist[key] == temp]
        mentor1 = random.choice(res)

        del submentorlist[mentor1]
        aggregatedmentorlist.append(mentor1)
        
        #once found, also update mentor view sheet with mentee name
        for mentor_name in range(2,mentor_view.max_row+1):
            if mentor_view.cell(mentor_name,1).value == mentor1:
                if(mentor_view.cell(mentor_name,2).value == None):
                    mentor_view.cell(mentor_name,2).value = matched_sheet.cell(mentee_name,2).value #R1 mentee name
                elif(mentor_view.cell(mentor_name,5).value == None):
                    mentor_view.cell(mentor_name,5).value = matched_sheet.cell(mentee_name,2).value #R1 mentee name
                elif(mentor_view.cell(mentor_name,6).value == None):
                    mentor_view.cell(mentor_name,6).value = matched_sheet.cell(mentee_name,2).value #R1 mentee name
                elif(mentor_view.cell(mentor_name,7).value == None):
                    mentor_view.cell(mentor_name,7).value = matched_sheet.cell(mentee_name,2).value #R1 mentee name
                else:
                    mentor_view.cell(mentor_name,8).value = matched_sheet.cell(mentee_name,2).value #R1 mentee name
        
        #update the mentor information in the matched sheet
        matched_sheet.cell(mentee_name,9).value = mentor1  #name of mentor
        for cc_name in range(2,mentors_sheet.max_row+1):   
            if mentors_sheet.cell(cc_name,5).value == mentor1:
                matched_sheet.cell(mentee_name,10).value = mentors_sheet.cell(cc_name,10).value #dept
                matched_sheet.cell(mentee_name,11).value = mentors_sheet.cell(cc_name,9).value #cohort    
        
        temp =  min(submentorlist.values())
        res = [key for key in submentorlist if submentorlist[key] == temp]
        mentor2 = random.choice(res)

        del submentorlist[mentor2]
        aggregatedmentorlist.append(mentor2)
        
        #once found, also update mentor view sheet with mentee name
        for mentor_name in range(2,mentor_view.max_row+1):
            if mentor_view.cell(mentor_name,1).value == mentor2:
                if(mentor_view.cell(mentor_name,3).value == None):
                    mentor_view.cell(mentor_name,3).value = matched_sheet.cell(mentee_name,2).value #R2 mentee name
                elif(mentor_view.cell(mentor_name,5).value == None):
                    mentor_view.cell(mentor_name,5).value = matched_sheet.cell(mentee_name,2).value #R2 mentee name
                elif(mentor_view.cell(mentor_name,6).value == None):
                    mentor_view.cell(mentor_name,6).value = matched_sheet.cell(mentee_name,2).value #R2 mentee name
                elif(mentor_view.cell(mentor_name,7).value == None):
                    mentor_view.cell(mentor_name,7).value = matched_sheet.cell(mentee_name,2).value #R2 mentee name
                else:
                    mentor_view.cell(mentor_name,8).value = matched_sheet.cell(mentee_name,2).value #R2 mentee name
        
        #update the mentor information in the matched sheet        
        matched_sheet.cell(mentee_name,12).value = mentor2  #name of mentor
        for cc_name in range(2,mentors_sheet.max_row+1):   
            if mentors_sheet.cell(cc_name,5).value == mentor2:
                matched_sheet.cell(mentee_name,13).value = mentors_sheet.cell(cc_name,10).value #dept
                matched_sheet.cell(mentee_name,14).value = mentors_sheet.cell(cc_name,9).value #cohort
              
        temp =  min(submentorlist.values())
        res = [key for key in submentorlist if submentorlist[key] == temp]
        mentor3 = random.choice(res)

        aggregatedmentorlist.append(mentor3)
        
        #once found, also update mentor view sheet with mentee name
        for mentor_name in range(2,mentor_view.max_row+1):
            if mentor_view.cell(mentor_name,1).value == mentor3:
                if(mentor_view.cell(mentor_name,4).value == None):
                    mentor_view.cell(mentor_name,4).value = matched_sheet.cell(mentee_name,2).value #R3 mentee name
                elif(mentor_view.cell(mentor_name,5).value == None):
                    mentor_view.cell(mentor_name,5).value = matched_sheet.cell(mentee_name,2).value #R3 mentee name
                elif(mentor_view.cell(mentor_name,6).value == None):
                    mentor_view.cell(mentor_name,6).value = matched_sheet.cell(mentee_name,2).value #R3 mentee name
                elif(mentor_view.cell(mentor_name,7).value == None):
                    mentor_view.cell(mentor_name,7).value = matched_sheet.cell(mentee_name,2).value #R3 mentee name
                else:
                    mentor_view.cell(mentor_name,8).value = matched_sheet.cell(mentee_name,2).value #R3 mentee name
                        
        
        #update the mentor information in the matched sheet        
        matched_sheet.cell(mentee_name,15).value = mentor3  #name of mentor
        for cc_name in range(2,mentors_sheet.max_row+1):   
            if mentors_sheet.cell(cc_name,5).value == mentor3:
                matched_sheet.cell(mentee_name,16).value = mentors_sheet.cell(cc_name,10).value #dept
                matched_sheet.cell(mentee_name,17).value = mentors_sheet.cell(cc_name,9).value #cohort
        
    #You can save the template as another file to create a new file here too.
    matched.save("matched_listc.xlsx")
    print("Range copied and pasted!")
    
#    print(Counter(aggregatedmentorlist))

createData()
######################################previous logic#####################################################             
        # start mentor 1 matching       
    #assign round 1 mentor to each mentee    
#    for mentee_name in range(2,matched_sheet.max_row+1):
#        #create filtered list of mentors for the mentee based on department, cohort, mentor availability
#        for mentor1_name in range(2,mentors_sheet.max_row+1):
#            #mentor and mentee should not belong to the same dept
#            if (matched_sheet.cell(mentee_name,5).value != mentors_sheet.cell(mentor1_name,10).value) and (matched_sheet.cell(mentee_name,6).value != mentors_sheet.cell(mentor1_name,9).value):
#                if (canMentorbeused(mentors_sheet.cell(mentor1_name,5).value) == 1) :
#                    pmentor1list.append(mentors_sheet.cell(mentor1_name,5).value)
#                    r1mentorlist.append(mentors_sheet.cell(mentor1_name,5).value)
#               
#            #choose a random mentor from the filtered list            
#            mentor1 = random.choice(pmentor1list)
#                
#        matched_sheet.cell(mentee_name,9).value = mentor1  #name of mentor
#        for cc_name in range(2,mentors_sheet.max_row+1):   
#            if mentors_sheet.cell(cc_name,5).value == mentor1:
#                matched_sheet.cell(mentee_name,10).value = mentors_sheet.cell(cc_name,10).value #dept
#                matched_sheet.cell(mentee_name,11).value = mentors_sheet.cell(cc_name,9).value #cohort         
             

     # start mentor 2 matching
#    #assign round 2 mentor to each mentee    
#    for mentee_name in range(2,matched_sheet.max_row+1):
#        #create filtered list of mentors for the mentee based on department, cohort, mentor availability
#        for mentor2_name in range(2,mentors_sheet.max_row+1):
#            #mentor and mentee should not belong to the same dept
#            if (matched_sheet.cell(mentee_name,5).value != mentors_sheet.cell(mentor2_name,10).value) and (matched_sheet.cell(mentee_name,6).value != mentors_sheet.cell(mentor2_name,9).value):
#                if (canMentorbeused(mentors_sheet.cell(mentor2_name,5).value) == 1):
#                    pmentor2list.append(mentors_sheet.cell(mentor2_name,5).value)                    
#                    break
#                
#                    #choose a random mentor from the filtered list           
#                    mentor2 = random.choice(pmentor2list)                   
#            break
#            r2mentorlist.append(mentor2)
#                
#                
#                
#        matched_sheet.cell(mentee_name,12).value = mentor2  #name of mentor
#        for cc_name in range(2,mentors_sheet.max_row+1):   
#            if mentors_sheet.cell(cc_name,5).value == mentor2:
#                matched_sheet.cell(mentee_name,13).value = mentors_sheet.cell(cc_name,10).value #dept
#                matched_sheet.cell(mentee_name,14).value = mentors_sheet.cell(cc_name,9).value #cohort

  # start mentor 3 matching       
    #assign round 3 mentor to each mentee    
#    for mentee_name in range(2,matched_sheet.max_row+1):
#        #create filtered list of mentors for the mentee based on department, cohort, mentor availability
#        for mentor3_name in range(2,mentors_sheet.max_row+1):
#            #mentor and mentee should not belong to the same dept
#            if (matched_sheet.cell(mentee_name,5).value != mentors_sheet.cell(mentor3_name,10).value) and (matched_sheet.cell(mentee_name,6).value != mentors_sheet.cell(mentor3_name,9).value):
#                if (canMentorbeused(mentors_sheet.cell(mentor3_name,5).value) == 1) :
#                    pmentor3list.append(mentors_sheet.cell(mentor3_name,5).value)
#                    r3mentorlist.append(mentors_sheet.cell(mentor3_name,5).value)
#               
#            #choose a random mentor from the filtered list            
#            mentor3 = random.choice(pmentor3list)
#                
#        matched_sheet.cell(mentee_name,12).value = mentor3  #name of mentor
#        for cc_name in range(2,mentors_sheet.max_row+1):   
#            if mentors_sheet.cell(cc_name,5).value == mentor3:
#                matched_sheet.cell(mentee_name,13).value = mentors_sheet.cell(cc_name,10).value #dept
#                matched_sheet.cell(mentee_name,14).value = mentors_sheet.cell(cc_name,9).value #cohort

              
    # start mentor 2 matching based on department
#    for mentee_name in range(2,matched_sheet.max_row+1):
#        for mentor2_name in range(2,mentors_sheet.max_row+1):            
#            if matched_sheet.cell(mentee_name,5).value != mentors_sheet.cell(mentor2_name,10).value:
#                pmentor2list.append(mentors_sheet.cell(mentor2_name,5).value) 
#        mentor2 = random.choice(pmentor2list)
#        matched_sheet.cell(mentee_name,12).value = mentor2  #name of mentor
#        for cc_name in range(2,mentors_sheet.max_row+1):            
#            if mentors_sheet.cell(cc_name,5).value == mentor2:
#                matched_sheet.cell(mentee_name,13).value = mentors_sheet.cell(cc_name,10).value #dept
#                matched_sheet.cell(mentee_name,14).value = mentors_sheet.cell(cc_name,9).value #cohort
#                
#                
#                
#                
#                
#    # start mentor 3 matching based on department
#    for mentee_name in range(2,matched_sheet.max_row+1):
##        print(pmentor1list)
#        for mentor3_name in range(2,mentors_sheet.max_row+1):            
#            if matched_sheet.cell(mentee_name,5).value != mentors_sheet.cell(mentor3_name,10).value:# and mentors_sheet.cell(mentor3_name,5).value not in pmentor1list:
#                pmentor3list.append(mentors_sheet.cell(mentor3_name,5).value)                
#        mentor3 = random.choice(pmentor3list)
#        matched_sheet.cell(mentee_name,15).value = mentor3  #name of mentor
#        for cc_name in range(2,mentors_sheet.max_row+1):            
#            if mentors_sheet.cell(cc_name,5).value == mentor3:
#                matched_sheet.cell(mentee_name,16).value = mentors_sheet.cell(cc_name,10).value #dept
#                matched_sheet.cell(mentee_name,17).value = mentors_sheet.cell(cc_name,9).value #cohort 
    
    #temporary code to add department info into mentee and mentor sheets from the CC list
#    for matched_name in range(2,mentees_sheet.max_row+1):
#        for cc_name in range(2,tees_sheet.max_row+1):            
#            if mentees_sheet.cell(matched_name,5).value == tees_sheet.cell(cc_name,3).value:
#                mentees_sheet.cell(matched_name,10).value = tees_sheet.cell(cc_name,9).value
#                
#        for cc_name in range(2,mentees_sheet.max_row+1):            
#            if mentees_sheet.cell(matched_name,5).value == edonly_sheet.cell(cc_name,3).value:
#                mentees_sheet.cell(matched_name,10).value = edonly_sheet.cell(cc_name,9).value
#    
#    mentees.save("mentees.xlsx")
#    
#    for matched_name in range(2,mentors_sheet.max_row+1):
#        for cc_name in range(2,tors_sheet.max_row+1):            
#            if mentors_sheet.cell(matched_name,5).value == tors_sheet.cell(cc_name,3).value:
#                mentors_sheet.cell(matched_name,10).value = tors_sheet.cell(cc_name,9).value
#                
#        for cc_name in range(2,mentees_sheet.max_row+1):            
#            if mentors_sheet.cell(matched_name,5).value == edonly_sheet.cell(cc_name,3).value:
#                mentors_sheet.cell(matched_name,10).value = edonly_sheet.cell(cc_name,9).value
#    
#    mentors.save("mentors.xlsx")
######################################previous logic#####################################################       
